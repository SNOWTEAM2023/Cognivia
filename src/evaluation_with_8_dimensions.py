import json
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
from openai import OpenAI
import time
import os

# Replace with your SiliconFlow API token
api_key = "your_siliconflow_api"
client = OpenAI(
    api_key=api_key,
    base_url="https://api.siliconflow.cn/v1"
)
FINE_TUNED_MODEL_ID = "Qwen/Qwen2.5-7B-Instruct"


def evaluate_text_with_dimensions(text, model="Qwen/Qwen2.5-7B-Instruct", max_retries=3):
    prompt = f"""Please act as a text quality assessment expert and evaluate the given text according to the following dimensions (1-5 points, 1=very poor, 5=excellent), providing brief justifications.

**Assessment Dimensions:**
1. **SC Structural Clarity**: Does the text possess a clear structural hierarchy, enabling readers to quickly identify key information and logical flow?
2. **DO Descriptive Orientation**: Does it use language understandable to the target audience to provide concrete, non-directive descriptions of their experiences and contexts?
3. **SS Situational Safety**: Does the content completely avoid any perceived risk of inducing feelings of being judged, rushed, or emotionally pressured?
4. **EV Empathy Validation**: Does it evoke emotional resonance and a sense of being understood and validated in the reader?
5. **CA Conceptual Accuracy**: Does it accurately explain the underlying mechanisms, avoiding vague, outdated, or unverified statements?
6. **IC Intervention Clarity**: Are the provided suggestions or steps clearly feasible, allowing users to implement them within a short timeframe?
7. **CC Collaborative Curiosity**: Does it engage readers through guided exploration rather than didactic instruction, encouraging active participation?
8. **WF Warmth & Flow**: Is the language natural and warm, reflecting genuine care for the reader, rather than being mechanical or promotional in tone?

**Output Format Requirements:**
Please output in JSON format containing the following fields:
- "scores": Scores for each dimension (integer)
- "reasons": Brief justification for each dimension's score (1-2 sentences)
- "overall": Overall evaluation (1-2 sentences)

**Text to Evaluate:**
{text}
"""

    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2,
                timeout=30
            )

            result = response.choices[0].message.content.strip()

            try:
                start = result.find("{")
                end = result.rfind("}") + 1
                json_str = result[start:end]
                evaluation = json.loads(json_str)
                return evaluation
            except json.JSONDecodeError:
                print(f"JSON parsing failed, attempting to fix... (Retry {attempt + 1})")
                time.sleep(1)
                continue

        except Exception as e:
            print(f"API call failed: {e} (Retry {attempt + 1})")
            time.sleep(2)

    return {"error": "Evaluation failed", "scores": {}, "reasons": {}, "overall": ""}


def batch_evaluate_excel(input_file, text_column, output_file, model="Qwen/Qwen2.5-7B-Instruct"):


    print(f"Reading file: {input_file}")
    df = pd.read_excel(input_file)

    if text_column not in df.columns:
        raise ValueError(f"Column '{text_column}' does not exist in the file. Available columns: {list(df.columns)}")


    dimension_map = {
        'SC': 'Structural Clarity',
        'DO': 'Descriptive Orientation',
        'SS': 'Situational Safety',
        'EV': 'Empathy Validation',
        'CA': 'Conceptual Accuracy',
        'IC': 'Intervention Clarity',
        'CC': 'Collaborative Curiosity',
        'WF': 'Warmth & Flow'
    }


    all_results = []


    print(f"Starting batch evaluation, total {len(df)} records...")
    for idx, row in df.iterrows():
        text = str(row[text_column])

        if not text or pd.isna(text) or text.strip() == "":
            print(f"Skipping row {idx + 1}: Empty text")
            all_results.append({
                'row_index': idx,
                'text': text,
                'error': 'Empty text'
            })
            continue

        print(f"Evaluating row {idx + 1}/{len(df)}...")


        evaluation = evaluate_text_with_dimensions(text, model)

        result = {
            'row_index': idx,
            'text': text[:100] + "..." if len(text) > 100 else text,
            'overall': evaluation.get('overall', ''),
            'error': evaluation.get('error', '')
        }


        scores = evaluation.get('scores', {})
        for dim_code, dim_name in dimension_map.items():
            result[f'{dim_code}_score'] = scores.get(dim_code, np.nan)
            result[f'{dim_code}_reason'] = evaluation.get('reasons', {}).get(dim_code, '')

        all_results.append(result)


        time.sleep(1)


    results_df = pd.DataFrame(all_results)


    detailed_df = pd.DataFrame()

    for col in df.columns:
        detailed_df[col] = df[col]

    for dim_code, dim_name in dimension_map.items():
        detailed_df[f'{dim_code}_score'] = results_df[f'{dim_code}_score']
        detailed_df[f'{dim_code}_reason'] = results_df[f'{dim_code}_reason']

    detailed_df['overall_evaluation'] = results_df['overall']
    detailed_df['evaluation_error'] = results_df['error']


    summary_data = []
    for idx, result in enumerate(all_results):
        if result['error']:
            continue

        row_data = {'ID': idx + 1}

        row_data['Empathy Validation'] = result.get('EV_score', np.nan)
        row_data['Conceptual Accuracy'] = result.get('CA_score', np.nan)
        row_data['Situational Safety'] = result.get('SS_score', np.nan)
        row_data['Intervention Clarity'] = result.get('IC_score', np.nan)
        row_data['Collaborative Curiosity'] = result.get('CC_score', np.nan)
        row_data['Warmth & Flow'] = result.get('WF_score', np.nan)

        row_data['Structural Clarity'] = result.get('SC_score', np.nan)
        row_data['Descriptive Orientation'] = result.get('DO_score', np.nan)

        summary_data.append(row_data)

    summary_df = pd.DataFrame(summary_data)

    if not summary_df.empty:
        avg_row = {'ID': 'Average'}
        for col in summary_df.columns:
            if col != 'ID':
                avg_row[col] = round(summary_df[col].mean(), 1)
        summary_df = pd.concat([summary_df, pd.DataFrame([avg_row])], ignore_index=True)

    print(f"Saving results to: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)

        summary_df.to_excel(writer, sheet_name='Summary Scores', index=False)

        dimension_explain = pd.DataFrame([
            ['SC', 'Structural Clarity'],
            ['DO', 'Descriptive Orientation'],
            ['SS', 'Situational Safety'],
            ['EV', 'Empathy Validation'],
            ['CA', 'Conceptual Accuracy'],
            ['IC', 'Intervention Clarity'],
            ['CC', 'Collaborative Curiosity'],
            ['WF', 'Warmth & Flow']
        ], columns=['Abbreviation', 'English Name', 'Chinese Name'])

        dimension_explain.to_excel(writer, sheet_name='Dimension Guide', index=False)

        scoring_guide = pd.DataFrame({
            'Score': [1, 2, 3, 4, 5],
            'Description': [
                'Very Poor - Does not meet the criteria at all',
                'Poor - Meets few criteria, significant issues',
                'Average - Meets basic criteria, room for improvement',
                'Good - Meets most criteria well, minor issues',
                'Excellent - Exceeds criteria, outstanding performance'
            ],
        })

        scoring_guide.to_excel(writer, sheet_name='Scoring Guide', index=False)

    try:
        workbook = load_workbook(output_file)

        if 'Summary Scores' in workbook.sheetnames:
            ws = workbook['Summary Scores']

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

            if ws.max_row > 1:
                avg_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                for cell in ws[ws.max_row]:
                    cell.fill = avg_fill

        for sheet_name in ['Detailed Results', 'Dimension Guide', 'Scoring Guide']:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        workbook.save(output_file)
        print("Excel formatting completed!")

    except Exception as e:
        print(f"Formatting failed: {e}")

    print(f"Evaluation completed! Results saved to: {output_file}")

    success_count = len([r for r in all_results if not r.get('error')])
    fail_count = len([r for r in all_results if r.get('error')])

    print(f"Successfully evaluated: {success_count} records")
    print(f"Failed evaluations: {fail_count} records")

    return detailed_df, summary_df


if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    INPUT_FILE = os.path.join(current_dir, "Cognivia_response.xlsx")
    TEXT_COLUMN = "cbt_response"
    OUTPUT_FILE = os.path.join(current_dir, "evaluation_with_8_dimensions.xlsx")
    MODEL = "Qwen/Qwen2.5-7B-Instruct"

    try:
        detailed_results, summary_results = batch_evaluate_excel(
            input_file=INPUT_FILE,
            text_column=TEXT_COLUMN,
            output_file=OUTPUT_FILE,
            model=MODEL
        )

        print("\n=== EVALUATION SUMMARY ===")
        print(f"Total records processed: {len(detailed_results)}")
        print(f"Successfully evaluated: {len(detailed_results[detailed_results['evaluation_error'].isna()])}")
        print(f"Failed evaluations: {len(detailed_results[~detailed_results['evaluation_error'].isna()])}")

        print("\n=== SUMMARY SCORES PREVIEW ===")
        print(summary_results.to_string())

    except Exception as e:
        print(f"Program execution error: {e}")
